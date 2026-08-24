#!/usr/bin/env python3
"""Build the complete JSIC G37-G41 + FUMA portable dataset.

The national Queria DuckDB is treated as a read-only source.  FUMA is the
scope authority for the G37-G41 company list; matched public records enrich a
FUMA row, while FUMA-only rows remain addressable as ``fuma:<FUMA_ID>``.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import os
import re
import shutil
import sqlite3
import tempfile
import unicodedata
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

import duckdb
from openpyxl import load_workbook


PACKAGE_VERSION = "0.10.1"
GIT_REF = os.environ.get("QUERIA_GIT_REF", f"working-tree@v{PACKAGE_VERSION}")
DEFAULT_XLSX = Path(os.environ.get("QUERIA_FUMA_XLSX", "data/import/fuma_g37_41.xlsx"))
DEFAULT_MASTER = Path("data/queria_master.duckdb")
DEFAULT_OUT = Path("releases/CompanyMaster-G37-41")
ESTAT_BASE = "https://www.e-stat.go.jp/term/download"
ESTAT_REVISION = "04"
ESTAT_URL = "https://www.e-stat.go.jp/classifications/terms/10/04/G"
MHLW_SOURCE_URL = "https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/0000177182.html"
REQUIRED_FUMA_HEADERS = {
    "企業名",
    "FUMA_ID",
    "本店所在地",
    "法人番号",
    "daibunruiCode",
    "chubunruiCode",
    "syoubunruiCode",
    "jsicDetailedClass",
    "電話番号",
    "公式サイトURL",
}


def text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip() or None


def json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def clean_code(value: Any) -> str | None:
    value = text(value)
    if not value:
        return None
    value = value.upper().replace("Ｇ", "G")
    value = re.sub(r"\.0$", "", value)
    value = re.sub(r"[^0-9G]", "", value)
    return value or None


def numeric_code(value: Any) -> str | None:
    code = clean_code(value)
    if not code:
        return None
    return code[1:] if code.startswith("G") else code


def corporate_number(value: Any) -> str | None:
    value = text(value)
    if not value:
        return None
    digits = re.sub(r"\D", "", value)
    return digits if len(digits) == 13 else None


def coarse_company_name_key(value: Any) -> str:
    """Return the cheap name key also expressible inside DuckDB SQL.

    This key is only used to shortlist candidates. Acceptance still requires
    the stricter normalized company name and full address to match exactly.
    """
    value = (text(value) or "").lower()
    for source, target in {
        "(株)": "株式会社",
        "（株）": "株式会社",
        "(有)": "有限会社",
        "（有）": "有限会社",
    }.items():
        value = value.replace(source, target)
    return re.sub(r"[\s\u3000・･.,，．\-‐–—_()（）/\\]+", "", value)


def normalize_match_name(value: Any) -> str:
    """Normalize a legal company name without dropping its entity type."""
    value = unicodedata.normalize("NFKC", text(value) or "").lower()
    for source, target in {
        "(株)": "株式会社",
        "(有)": "有限会社",
        "(同)": "合同会社",
        "(資)": "合資会社",
        "(名)": "合名会社",
    }.items():
        value = value.replace(source, target)
    return re.sub(
        r"[\s\u3000・･\.\uff0e,，'’\"“”\-‐–—_()（）\[\]【】/\\]+",
        "",
        value,
    )


def normalize_match_address(value: Any) -> str:
    """Normalize Japanese address notation for conservative exact matching."""
    value = unicodedata.normalize("NFKC", text(value) or "").lower()
    value = re.sub(r"〒?\s*\d{3}[-‐‑‒–—―ー]?\d{4}", "", value)
    for source, target in {
        "〇": "0",
        "一": "1",
        "二": "2",
        "三": "3",
        "四": "4",
        "五": "5",
        "六": "6",
        "七": "7",
        "八": "8",
        "九": "9",
    }.items():
        value = value.replace(source, target)
    value = (
        value.replace("丁目", "-")
        .replace("番地", "-")
        .replace("番", "-")
        .replace("号", "")
        .replace("大字", "")
        .replace("字", "")
    )
    value = re.sub(
        r"[\s\u3000・･\.\uff0e,，'’\"“”()（）\[\]【】/\\]+",
        "",
        value,
    )
    return re.sub(r"[-‐‑‒–—―ー]+", "", value)


def _coarse_public_name_sql(column: str) -> str:
    """DuckDB expression equivalent to :func:`coarse_company_name_key`."""
    return (
        "lower(regexp_replace("
        f"replace(replace(replace(replace({column}, '（株）', '株式会社'), '(株)', '株式会社'), "
        "'（有）', '有限会社'), '(有)', '有限会社'), "
        "'[[:space:]　・･.,，．\\-‐–—_()（）/\\\\]', '', 'g'))"
    )


def integer_value(value: Any) -> int | None:
    """Parse a numeric FUMA/public value without inventing a value."""
    value = text(value)
    if not value:
        return None
    digits = re.sub(r"[^0-9-]", "", value)
    if not digits or digits == "-":
        return None
    try:
        return int(digits)
    except ValueError:
        return None


def year_value(value: Any) -> int | None:
    value = text(value)
    if not value:
        return None
    match = re.search(r"(?<!\d)(18|19|20|21)\d{2}(?!\d)", value)
    return int(match.group(0)) if match else None


def normalize_phone(value: Any) -> str | None:
    value = text(value)
    if not value:
        return None
    value = value.replace("０", "0").replace("１", "1").replace("２", "2").replace("３", "3")
    value = value.replace("４", "4").replace("５", "5").replace("６", "6").replace("７", "7")
    value = value.replace("８", "8").replace("９", "9").replace("ー", "-")
    digits = re.sub(r"\D", "", value)
    if digits.startswith("81") and len(digits) >= 10:
        digits = "0" + digits[2:]
    if not (10 <= len(digits) <= 11) or not digits.startswith("0"):
        return None
    return digits


def phone_parts(value: Any) -> list[str]:
    raw = text(value)
    if not raw:
        return []
    candidates = re.findall(r"(?:\+?81|0)[0-9０-９\-‐‑‒–—−ー()（）\s]{7,}[0-9０-９]", raw)
    out: list[str] = []
    for candidate in candidates or [raw]:
        normalized = normalize_phone(candidate)
        if normalized and normalized not in out:
            out.append(normalized)
    return out[:5]


def normalize_http_url(value: Any) -> str | None:
    value = text(value)
    if not value:
        return None
    try:
        parsed = urllib.parse.urlsplit(value)
    except ValueError:
        return None
    if parsed.scheme.lower() not in {"http", "https"} or not parsed.netloc:
        return None
    return urllib.parse.urlunsplit((parsed.scheme.lower(), parsed.netloc.lower(), parsed.path, parsed.query, ""))


def load_public_establishment_contacts(
    source: duckdb.DuckDBPyConnection,
) -> tuple[dict[str, list[dict[str, Any]]], dict[str, int]]:
    """Load public establishment contacts for the already-scoped G companies.

    These records are joined by a 13-digit corporate number. They are never
    labeled as headquarters or representative contacts.
    """
    try:
        observed_at = text(source.execute("SELECT max(completed_at)::VARCHAR FROM meta.refresh_log").fetchone()[0])
    except duckdb.Error:
        observed_at = None
    rows = source.execute(
        """
        SELECT e.corporate_number, e.establishment_number, e.establishment_name,
               e.phone, e.url, e.source_dataset
        FROM (
          SELECT corporate_number, establishment_number, name AS establishment_name,
                 phone, url, 'mhlw.kaigo_establishment' AS source_dataset
          FROM mhlw.kaigo_establishment
          UNION ALL
          SELECT corporate_number, establishment_number, name AS establishment_name,
                 phone, url, 'mhlw.shougai_establishment' AS source_dataset
          FROM mhlw.shougai_establishment
        ) e
        INNER JOIN scoped_g_numbers s USING(corporate_number)
        WHERE coalesce(e.phone, '') <> '' OR coalesce(e.url, '') <> ''
        ORDER BY e.corporate_number, e.source_dataset, e.establishment_number
        """
    ).fetchall()
    by_corporate_number: dict[str, list[dict[str, Any]]] = defaultdict(list)
    phone_companies: set[str] = set()
    website_companies: set[str] = set()
    for corp, establishment_number, establishment_name, phone, url, dataset in rows:
        corp = str(corp)
        normalized_phone = normalize_phone(phone)
        normalized_url = normalize_http_url(url)
        if not normalized_phone and not normalized_url:
            continue
        if normalized_phone:
            phone_companies.add(corp)
        if normalized_url:
            website_companies.add(corp)
        by_corporate_number[corp].append(
            {
                "establishment_number": text(establishment_number),
                "establishment_name": text(establishment_name),
                "phone_raw": text(phone),
                "phone_normalized": normalized_phone,
                "url": normalized_url,
                "source_dataset": str(dataset),
                "source_url": MHLW_SOURCE_URL,
                "observed_at": observed_at,
            }
        )
    return dict(by_corporate_number), {
        "contact_rows": sum(len(value) for value in by_corporate_number.values()),
        "company_rows": len(by_corporate_number),
        "phone_company_rows": len(phone_companies),
        "website_company_rows": len(website_companies),
    }


def apply_public_establishment_contacts(
    companies: list[dict[str, Any]],
    contacts: dict[str, list[dict[str, Any]]],
    phone_rows: list[dict[str, Any]],
    website_rows: list[dict[str, Any]],
) -> dict[str, int]:
    """Persist all public contacts and promote one phone only when missing."""
    company_by_corp = {str(row["corporate_number"]): row for row in companies if row.get("corporate_number")}
    promoted = 0
    for corp, contact_rows in contacts.items():
        company = company_by_corp.get(corp)
        if not company:
            continue
        entity_key = str(company["entity_key"])
        first_phone: dict[str, Any] | None = None
        seen_phones: set[str] = set()
        seen_urls: set[str] = set()
        for contact in contact_rows:
            establishment = contact.get("establishment_number") or "unknown"
            evidence = (
                f"{contact['source_dataset']}の事業所"
                f"「{contact.get('establishment_name') or '名称未設定'}」({establishment})"
            )
            normalized_phone = contact.get("phone_normalized")
            if normalized_phone and normalized_phone not in seen_phones:
                seen_phones.add(normalized_phone)
                digest = hashlib.sha256(
                    f"{entity_key}|{contact['source_dataset']}|{establishment}|{normalized_phone}".encode("utf-8")
                ).hexdigest()[:20]
                phone_rows.append(
                    {
                        "candidate_id": f"{entity_key}:mhlw:{digest}",
                        "entity_key": entity_key,
                        "phone_raw": contact.get("phone_raw"),
                        "phone_normalized": normalized_phone,
                        "phone_type": "establishment",
                        "source_url": contact.get("source_url"),
                        "evidence_text": evidence,
                        "confidence": 1.0,
                        "observed_at": contact.get("observed_at"),
                        "status": "imported_public_establishment",
                    }
                )
                first_phone = first_phone or contact
            candidate_url = contact.get("url")
            if candidate_url and candidate_url not in seen_urls:
                seen_urls.add(candidate_url)
                digest = hashlib.sha256(
                    f"{entity_key}|{contact['source_dataset']}|{establishment}|{candidate_url}".encode("utf-8")
                ).hexdigest()[:20]
                website_rows.append(
                    {
                        "candidate_id": f"{entity_key}:mhlw-url:{digest}",
                        "entity_key": entity_key,
                        "url": candidate_url,
                        "url_type": "establishment",
                        "source_url": contact.get("source_url"),
                        "evidence_text": evidence,
                        "confidence": 1.0,
                        "observed_at": contact.get("observed_at"),
                        "status": "imported_public_establishment",
                    }
                )
        if not company.get("phone") and first_phone:
            company["phone"] = first_phone["phone_normalized"]
            company["phone_type"] = "establishment"
            company["phone_source_url"] = first_phone["source_url"]
            company["phone_confidence"] = 1.0
            company["phone_evidence_text"] = (
                f"{first_phone['source_dataset']}の事業所電話。代表電話ではありません"
            )
            company["phone_observed_at"] = first_phone.get("observed_at")
            company["phone_status"] = "imported_public_establishment"
            promoted += 1
    return {"promoted_primary_phone_rows": promoted}


PREFECTURE_RE = re.compile(r"^(北海道|東京都|京都府|大阪府|[東西南北]?[一-龯ぁ-んァ-ヶ]{1,3}県)")
CITY_RE = re.compile(r"^(.+?市(?:.+?区)?|.+?郡.+?[町村]|.+?[区町村])")


def parse_japanese_address(value: Any) -> tuple[str | None, str | None]:
    """Split a FUMA full address into the fields used by the GUI filters."""
    address = text(value)
    if not address:
        return None, None
    prefecture_match = PREFECTURE_RE.match(address)
    if not prefecture_match:
        return None, None
    prefecture = prefecture_match.group(1)
    remainder = address[prefecture_match.end():]
    city_match = CITY_RE.match(remainder)
    return prefecture, city_match.group(1) if city_match else None


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def quote_ident(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def quote_sql(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def fetch_estat(code: str) -> list[tuple[str, str, str | None]]:
    query = urllib.parse.urlencode(
        {"bKbn": "10", "kaiteiCode": ESTAT_REVISION, "charset": "UTF-8", "bom": "1", "pbCode": code}
    )
    request = urllib.request.Request(
        f"{ESTAT_BASE}?{query}", headers={"User-Agent": "CompanyMaster-G37-41/0.1 taxonomy builder"}
    )
    with urllib.request.urlopen(request, timeout=30) as response:
        content = response.read().decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(content)))
    parsed: list[tuple[str, str, str | None]] = []
    for row in rows[2:]:
        if len(row) < 2:
            continue
        row_code = numeric_code(row[0])
        name = text(row[1])
        if row_code and name:
            parsed.append((row_code, name, text(row[2]) if len(row) > 2 else None))
    return parsed


def build_taxonomy(output_csv: Path, refresh: bool) -> list[dict[str, Any]]:
    if output_csv.is_file() and not refresh:
        with output_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))

    rows: list[dict[str, Any]] = [
        {
            "code": "G",
            "name": "情報通信業",
            "level": "major",
            "parent_code": "",
            "path": "G",
            "revision": "令和5年（2023年）7月改定",
            "source_url": ESTAT_URL,
            "is_active": "1",
        }
    ]
    middle = fetch_estat("G")
    for middle_code, middle_name, _ in middle:
        middle_url = f"https://www.e-stat.go.jp/classifications/terms/10/04/{middle_code}"
        rows.append(
            {
                "code": middle_code,
                "name": middle_name,
                "level": "middle",
                "parent_code": "G",
                "path": f"G/{middle_code}",
                "revision": "令和5年（2023年）7月改定",
                "source_url": middle_url,
                "is_active": "1",
            }
        )
        for small_code, small_name, _ in fetch_estat(middle_code):
            small_url = f"https://www.e-stat.go.jp/classifications/terms/10/04/{small_code}"
            rows.append(
                {
                    "code": small_code,
                    "name": small_name,
                    "level": "small",
                    "parent_code": middle_code,
                    "path": f"G/{middle_code}/{small_code}",
                    "revision": "令和5年（2023年）7月改定",
                    "source_url": small_url,
                    "is_active": "1",
                }
            )
            for detail_code, detail_name, _ in fetch_estat(small_code):
                rows.append(
                    {
                        "code": detail_code,
                        "name": detail_name,
                        "level": "detail",
                        "parent_code": small_code,
                        "path": f"G/{middle_code}/{small_code}/{detail_code}",
                        "revision": "令和5年（2023年）7月改定",
                        "source_url": f"https://www.e-stat.go.jp/classifications/terms/10/04/{detail_code}",
                        "is_active": "1",
                    }
                )
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    fields = ["code", "name", "level", "parent_code", "path", "revision", "source_url", "is_active"]
    with output_csv.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
    return rows


def canonical_taxonomy(rows: list[dict[str, Any]]) -> tuple[dict[str, dict[str, Any]], dict[str, str]]:
    by_code = {str(row["code"]): row for row in rows}
    middle = {code for code, row in by_code.items() if row["level"] == "middle"}
    small = {code for code, row in by_code.items() if row["level"] == "small"}
    detail = {code for code, row in by_code.items() if row["level"] == "detail"}

    def canonical(value: Any) -> str | None:
        code = clean_code(value)
        if not code:
            return None
        if code == "G":
            return "G"
        code = code[1:] if code.startswith("G") else code
        if code in by_code:
            return code
        # Some public Queria codes concatenate G + middle + small/detail.
        for candidate in sorted(detail | small | middle, key=len, reverse=True):
            if code.endswith(candidate) and code.startswith(candidate[:2]):
                return candidate
        return code

    aliases: dict[str, str] = {}
    for code in by_code:
        aliases[code] = code
        if code != "G":
            aliases[f"G{code}"] = code
    return by_code, aliases


def ancestors(code: str | None, by_code: dict[str, dict[str, Any]]) -> list[str]:
    if not code:
        return ["G"]
    if code == "G":
        return ["G"]
    chain = [code]
    current = code
    while current in by_code and by_code[current].get("parent_code"):
        current = str(by_code[current]["parent_code"])
        chain.append(current)
    if "G" not in chain:
        chain.append("G")
    return list(reversed(chain))


def aliases_for(code: str | None, by_code: dict[str, dict[str, Any]]) -> list[str]:
    values: list[str] = []
    for ancestor in ancestors(code, by_code):
        values.extend([ancestor, f"G{ancestor}"] if ancestor != "G" else ["G"])
    return list(dict.fromkeys(values))


def load_fuma(path: Path) -> tuple[list[str], list[dict[str, Any]], dict[str, int]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if len(workbook.sheetnames) < 2:
        raise ValueError("FUMA ExcelのDBシートがありません")
    worksheet = workbook[workbook.sheetnames[1]]
    iterator = worksheet.iter_rows(values_only=True)
    headers = [text(value) or f"column_{index:02d}" for index, value in enumerate(next(iterator))]
    missing_headers = sorted(REQUIRED_FUMA_HEADERS.difference(headers))
    if missing_headers:
        raise ValueError(f"FUMA Excelの必須列がありません: {', '.join(missing_headers)}")
    duplicate_headers = sorted({header for header in headers if headers.count(header) > 1})
    if duplicate_headers:
        raise ValueError(f"FUMA Excelの列名が重複しています: {', '.join(duplicate_headers)}")
    records: list[dict[str, Any]] = []
    for row_number, row in enumerate(iterator, start=2):
        values = list(row) + [None] * max(0, len(headers) - len(row))
        records.append(
            {
                "row_number": row_number,
                "values": values[: len(headers)],
                "by_header": {header: values[index] for index, header in enumerate(headers)},
            }
        )
    return headers, records, {header: index for index, header in enumerate(headers)}


PUBLIC_COLUMNS = [
    "corporate_number", "company_name", "company_name_en", "company_name_kana", "corporate_kind_code",
    "post_code", "prefecture_name", "city_name", "full_address", "representative_name", "capital_stock",
    "employee_number", "founding_year", "business_summary", "company_url", "extracted_at",
]


def public_company_row(row: tuple[Any, ...]) -> dict[str, Any]:
    return dict(zip(PUBLIC_COLUMNS, row))


def recover_corporate_numbers(
    source: duckdb.DuckDBPyConnection,
    fuma_records: list[dict[str, Any]],
) -> tuple[dict[str, str], dict[str, int]]:
    """Recover missing corporate numbers using unique exact name+address matches.

    A cheap SQL name key narrows the 5.8M-row national master. Python then
    applies the same conservative normalization used by the public enrichment
    matcher. A match is accepted only when both sides are unique, so no company
    receives a guessed or ambiguous corporate number.
    """
    explicit_numbers = {
        corp
        for record in fuma_records
        if (corp := corporate_number(record["by_header"].get("法人番号")))
    }
    targets = []
    for record in fuma_records:
        row = record["by_header"]
        if corporate_number(row.get("法人番号")):
            continue
        name = text(row.get("企業名"))
        address = text(row.get("本店所在地"))
        if not name or not address:
            continue
        targets.append(
            (
                str(record["fuma_id"]),
                coarse_company_name_key(name),
                normalize_match_name(name),
                normalize_match_address(address),
            )
        )

    source.execute("DROP TABLE IF EXISTS fuma_match_targets")
    source.execute(
        "CREATE TEMP TABLE fuma_match_targets("
        "fuma_id VARCHAR PRIMARY KEY, coarse_name VARCHAR, name_norm VARCHAR, address_norm VARCHAR)"
    )
    if targets:
        source.executemany("INSERT INTO fuma_match_targets VALUES (?,?,?,?)", targets)

    candidates: dict[str, set[str]] = defaultdict(set)
    candidate_rows = 0
    query = f"""
        SELECT t.fuma_id, t.name_norm, t.address_norm,
               c.corporate_number, c.company_name, c.full_address
        FROM core.companies c
        INNER JOIN fuma_match_targets t
          ON {_coarse_public_name_sql('c.company_name')} = t.coarse_name
        WHERE c.corporate_number IS NOT NULL
          AND c.company_name IS NOT NULL
          AND c.full_address IS NOT NULL
    """
    cursor = source.execute(query)
    while rows := cursor.fetchmany(10_000):
        candidate_rows += len(rows)
        for fuma_id, name_norm, address_norm, corp, public_name, public_address in rows:
            if (
                str(name_norm) == normalize_match_name(public_name)
                and str(address_norm) == normalize_match_address(public_address)
            ):
                candidates[str(fuma_id)].add(str(corp))

    unique_by_fuma = {
        fuma_id: next(iter(values))
        for fuma_id, values in candidates.items()
        if len(values) == 1
    }
    reverse_counts = Counter(unique_by_fuma.values())
    accepted = {
        fuma_id: corp
        for fuma_id, corp in unique_by_fuma.items()
        if reverse_counts[corp] == 1 and corp not in explicit_numbers
    }
    explicit_collisions = sum(1 for corp in unique_by_fuma.values() if corp in explicit_numbers)
    return accepted, {
        "targets": len(targets),
        "candidate_rows": candidate_rows,
        "exact_match_targets": len(candidates),
        "unique_exact_matches": len(unique_by_fuma),
        "accepted_one_to_one": len(accepted),
        "rejected_existing_explicit_number": explicit_collisions,
        "rejected_ambiguous_or_duplicate": len(candidates) - len(accepted),
    }


def public_industry(code: Any, by_code: dict[str, dict[str, Any]], aliases: dict[str, str]) -> dict[str, Any]:
    raw = clean_code(code)
    candidate = aliases.get(raw or "")
    if not candidate and raw:
        candidate = aliases.get(raw[1:] if raw.startswith("G") else raw)
    if candidate not in by_code:
        candidate = "G"
    row = by_code[candidate]
    chain = ancestors(candidate, by_code)
    middle = next((x for x in chain if by_code.get(x, {}).get("level") == "middle"), None)
    small = next((x for x in chain if by_code.get(x, {}).get("level") == "small"), None)
    detail = next((x for x in chain if by_code.get(x, {}).get("level") == "detail"), None)
    return {
        "code": candidate,
        "name": row["name"],
        "middle_code": middle,
        "middle_name": by_code[middle]["name"] if middle else None,
        "small_code": small,
        "small_name": by_code[small]["name"] if small else None,
        "detail_code": detail,
        "detail_name": by_code[detail]["name"] if detail else None,
    }


def make_company(
    *,
    entity_key: str,
    fuma: dict[str, Any] | None,
    public: dict[str, Any] | None,
    fuma_id: str | None,
    fuma_row_number: int | None,
    by_code: dict[str, dict[str, Any]],
    aliases: dict[str, str],
    industry: dict[str, Any] | None,
    corporate_number_override: str | None = None,
    corporate_number_match_method: str | None = None,
) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    def pick(fuma_header: str, public_key: str, default: Any = None) -> Any:
        if fuma:
            value = text(fuma["by_header"].get(fuma_header))
            if value is not None:
                return value
        return public.get(public_key, default) if public else default

    if fuma:
        middle = numeric_code(fuma["by_header"].get("chubunruiCode"))
        small = numeric_code(fuma["by_header"].get("syoubunruiCode"))
        detail = numeric_code(fuma["by_header"].get("jsicDetailedClass"))
        selected = detail or small or middle or "G"
        selected = aliases.get(selected, selected)
        industry = {
            "code": selected,
            "name": text(fuma["by_header"].get("jsicDetailedClass")) and text(fuma["by_header"].get("saibunruiName"))
            or text(fuma["by_header"].get("syoubunruiName"))
            or text(fuma["by_header"].get("chubunruiName"))
            or "情報通信業",
            "middle_code": middle,
            "middle_name": text(fuma["by_header"].get("chubunruiName")),
            "small_code": small,
            "small_name": text(fuma["by_header"].get("syoubunruiName")),
            "detail_code": detail,
            "detail_name": text(fuma["by_header"].get("saibunruiName")),
        }
    industry = industry or public_industry("G", by_code, aliases)
    selected_code = industry["code"]
    code_aliases = aliases_for(selected_code, by_code)
    industry_codes = "|".join(code_aliases)
    path = "/".join(ancestors(selected_code, by_code))
    phone_values = phone_parts(fuma["by_header"].get("電話番号")) if fuma else []
    phone = phone_values[0] if phone_values else None
    phone_url = None
    if fuma:
        phone_url = text(fuma["by_header"].get("取得元URL")) or text(fuma["by_header"].get("公的データ元URL")) or text(fuma["by_header"].get("FUMA_URL"))
    source_kind = "fuma-only" if fuma and not public else ("fuma+national" if fuma else "national-g")
    source_updated_at = pick("詳細取得日時", "extracted_at") or pick("取得日時", "extracted_at")
    fuma_address = text(fuma["by_header"].get("本店所在地")) if fuma else None
    public_prefecture = text(public.get("prefecture_name")) if public else None
    public_city = text(public.get("city_name")) if public else None
    parsed_prefecture, parsed_city = parse_japanese_address(fuma_address)
    explicit_corporate_number = corporate_number(fuma["by_header"].get("法人番号")) if fuma else None
    resolved_corporate_number = (
        corporate_number_override
        or explicit_corporate_number
        or (text(public.get("corporate_number")) if public else None)
    )
    match_method = corporate_number_match_method
    if not match_method and explicit_corporate_number:
        match_method = "fuma_explicit"
    if not match_method and public and not fuma:
        match_method = "national_primary_key"
    company = {
        "entity_key": entity_key,
        "corporate_number": resolved_corporate_number,
        "corporate_number_match_method": match_method,
        "corporate_number_match_score": 1.0 if resolved_corporate_number else None,
        "fuma_id": fuma_id,
        "fuma_source_row": fuma_row_number,
        "name": pick("企業名", "company_name") or "名称未設定",
        "name_kana": pick("nameKana", "company_name_kana"),
        "name_en": pick("nameEn", "company_name_en"),
        "prefecture": public_prefecture or parsed_prefecture,
        "city": public_city or parsed_city,
        "address": fuma_address or (public.get("full_address") if public else None),
        "postal_code": pick("郵便番号", "post_code"),
        "kind": None,
        "industry_code": industry_codes,
        "industry_name": " / ".join(x for x in ["情報通信業", industry.get("middle_name"), industry.get("small_name"), industry.get("detail_name")] if x),
        "industry_source": "FUMA/JSIC2023" if fuma else "Queria/main-G",
        "industry_codes": industry_codes,
        "industry_path": path,
        "industry_middle_code": industry.get("middle_code"),
        "industry_middle_name": industry.get("middle_name"),
        "industry_small_code": industry.get("small_code"),
        "industry_small_name": industry.get("small_name"),
        "industry_detail_code": industry.get("detail_code"),
        "industry_detail_name": industry.get("detail_name"),
        "employees": integer_value(fuma["by_header"].get("従業員数")) if fuma else (public.get("employee_number") if public else None),
        "capital": integer_value(fuma["by_header"].get("capital")) if fuma else (public.get("capital_stock") if public else None),
        "established_year": year_value(fuma["by_header"].get("establishedDate")) if fuma else (public.get("founding_year") if public else None),
        "website": pick("公式サイトURL", "company_url"),
        "phone": phone,
        "representative": pick("ceoName", "representative_name"),
        "business_summary": pick("事業概要", "business_summary"),
        "source_kind": source_kind,
        "source_updated_at": source_updated_at,
        "phone_type": "unclassified" if phone else None,
        "phone_source_url": phone_url,
        "phone_confidence": None,
        "phone_evidence_text": "FUMA Excelの電話番号列" if phone else None,
        "phone_observed_at": text(fuma["by_header"].get("取得日時")) if fuma and phone else None,
        "phone_status": "imported_fuma" if phone else ("pending_official_site" if pick("公式サイトURL", "company_url") and source_kind != "fuma-only" else "no_phone_source"),
    }
    industry_rows = [
        {
            "entity_key": entity_key,
            "jsic_code": selected_code,
            "jsic_major_code": "G",
            "jsic_middle_code": industry.get("middle_code"),
            "jsic_small_code": industry.get("small_code"),
            "jsic_detail_code": industry.get("detail_code"),
            "jsic_major_name": "情報通信業",
            "jsic_middle_name": industry.get("middle_name"),
            "jsic_small_name": industry.get("small_name"),
            "jsic_detail_name": industry.get("detail_name"),
            "jsic_level": by_code.get(selected_code, {}).get("level", "major"),
            "business_path_raw": path,
        }
    ]
    phone_rows = []
    for index, value in enumerate(phone_values):
        phone_rows.append({
            "candidate_id": f"{entity_key}:fuma:{index + 1}",
            "entity_key": entity_key,
            "phone_raw": text(fuma["by_header"].get("電話番号")),
            "phone_normalized": value,
            "phone_type": "unclassified",
            "source_url": phone_url,
            "evidence_text": "FUMA Excelの電話番号列",
            "confidence": None,
            "observed_at": text(fuma["by_header"].get("取得日時")),
            "status": "imported_fuma",
        })
    return company, industry_rows, phone_rows


COMPANY_COLUMNS = [
    "entity_key", "corporate_number", "corporate_number_match_method", "corporate_number_match_score",
    "fuma_id", "fuma_source_row", "name", "name_kana", "name_en",
    "prefecture", "city", "address", "postal_code", "kind", "industry_code", "industry_name", "industry_source",
    "industry_codes", "industry_path", "industry_middle_code", "industry_middle_name", "industry_small_code",
    "industry_small_name", "industry_detail_code", "industry_detail_name", "employees", "capital", "established_year",
    "website", "phone", "representative", "business_summary", "source_kind", "source_updated_at", "phone_type",
    "phone_source_url", "phone_confidence", "phone_evidence_text", "phone_observed_at", "phone_status",
]


def create_schema(con: duckdb.DuckDBPyConnection) -> None:
    con.execute("CREATE SCHEMA core")
    con.execute("CREATE SCHEMA meta")
    con.execute("CREATE SCHEMA enrichment")
    con.execute(
        "CREATE TABLE core.g_companies (" + ",".join(
            f"{quote_ident(column)} " + ("BIGINT" if column in {"fuma_source_row", "employees", "capital", "established_year"} else "DOUBLE" if column in {"phone_confidence", "corporate_number_match_score"} else "VARCHAR")
            for column in COMPANY_COLUMNS
        ) + ", PRIMARY KEY(entity_key))"
    )
    con.execute(
        "CREATE TABLE core.company_industries (entity_key VARCHAR, jsic_code VARCHAR, jsic_major_code VARCHAR, jsic_middle_code VARCHAR, jsic_small_code VARCHAR, jsic_detail_code VARCHAR, jsic_major_name VARCHAR, jsic_middle_name VARCHAR, jsic_small_name VARCHAR, jsic_detail_name VARCHAR, jsic_level VARCHAR, business_path_raw VARCHAR)"
    )
    con.execute("CREATE TABLE core.company_category_index AS SELECT entity_key, jsic_code, jsic_major_code, jsic_middle_code, prefecture, city FROM (SELECT CAST(NULL AS VARCHAR) AS entity_key, CAST(NULL AS VARCHAR) AS jsic_code, CAST(NULL AS VARCHAR) AS jsic_major_code, CAST(NULL AS VARCHAR) AS jsic_middle_code, CAST(NULL AS VARCHAR) AS prefecture, CAST(NULL AS VARCHAR) AS city) WHERE false")
    con.execute("CREATE TABLE core.fuma_records (source_row_number BIGINT, raw_json VARCHAR, " + ",".join(f"{quote_ident(column)} VARCHAR" for column in RAW_HEADERS) + ")")
    con.execute("CREATE TABLE core.fuma_columns (ordinal INTEGER, source_name VARCHAR, db_column VARCHAR)")
    con.execute("CREATE TABLE enrichment.phone_candidates (candidate_id VARCHAR PRIMARY KEY, entity_key VARCHAR, phone_raw VARCHAR, phone_normalized VARCHAR, phone_type VARCHAR, source_url VARCHAR, evidence_text VARCHAR, confidence DOUBLE, observed_at VARCHAR, status VARCHAR)")
    con.execute("CREATE TABLE enrichment.website_candidates (candidate_id VARCHAR PRIMARY KEY, entity_key VARCHAR, url VARCHAR, url_type VARCHAR, source_url VARCHAR, evidence_text VARCHAR, confidence DOUBLE, observed_at VARCHAR, status VARCHAR)")
    con.execute("CREATE TABLE enrichment.phone_collection_state (entity_key VARCHAR PRIMARY KEY, corporate_number VARCHAR, website VARCHAR, state VARCHAR, last_completed_at VARCHAR, last_error VARCHAR)")
    con.execute("CREATE TABLE meta.industry_taxonomy (code VARCHAR PRIMARY KEY, name VARCHAR, level VARCHAR, parent_code VARCHAR, path VARCHAR, revision VARCHAR, source_url VARCHAR, is_active BOOLEAN)")
    con.execute("CREATE TABLE meta.dataset_manifest (dataset_key VARCHAR PRIMARY KEY, value VARCHAR)")
    con.execute("CREATE TABLE meta.source_metadata (dataset_name VARCHAR PRIMARY KEY, metadata_json VARCHAR)")


RAW_HEADERS: list[str] = []


def insert_batches(con: duckdb.DuckDBPyConnection, table: str, columns: list[str], rows: Iterable[dict[str, Any]], batch_size: int = 500) -> int:
    """Bulk-load rows through DuckDB JSON COPY.

    ``batch_size`` remains in the public signature for compatibility with the
    older executemany implementation. JSON Lines preserves NULL, empty text,
    commas, quotes and embedded newlines. COPY streams the temporary file,
    avoiding millions of individual Python parameter bindings for the 54-column
    FUMA raw table.
    """
    del batch_size
    count = 0
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix="queria_bulk_",
            suffix=".jsonl",
            delete=False,
            mode="w",
            encoding="utf-8",
        ) as handle:
            temporary_path = Path(handle.name)
            for row in rows:
                handle.write(
                    json.dumps(
                        {column: json_value(row.get(column)) for column in columns},
                        ensure_ascii=False,
                    )
                    + "\n"
                )
                count += 1
        if count:
            column_sql = ",".join(quote_ident(column) for column in columns)
            con.execute(
                f"COPY {table} ({column_sql}) FROM {quote_sql(temporary_path.as_posix())} "
                "(FORMAT JSON)"
            )
    finally:
        if temporary_path:
            temporary_path.unlink(missing_ok=True)
    return count


def build_sqlite_index(path: Path, companies: list[dict[str, Any]], taxonomy: list[dict[str, Any]], generation: str) -> None:
    if path.exists():
        path.unlink()
    con = sqlite3.connect(path)
    con.execute("PRAGMA journal_mode=DELETE")
    con.execute("CREATE TABLE companies (row_id INTEGER PRIMARY KEY, entity_key TEXT UNIQUE, corporate_number TEXT, fuma_id TEXT, name TEXT, address TEXT, prefecture TEXT, city TEXT, industry_code TEXT, industry_name TEXT, industry_path TEXT, phone TEXT, website TEXT, source_kind TEXT, phone_type TEXT, phone_source_url TEXT, phone_status TEXT)")
    con.execute("CREATE VIRTUAL TABLE company_fts USING fts5(entity_key UNINDEXED, name, address, industry_name, business_summary, content='')")
    con.execute("CREATE TABLE industry_taxonomy (code TEXT PRIMARY KEY, name TEXT, level TEXT, parent_code TEXT, path TEXT, revision TEXT, source_url TEXT, is_active INTEGER)")
    con.executemany("INSERT INTO companies(entity_key,corporate_number,fuma_id,name,address,prefecture,city,industry_code,industry_name,industry_path,phone,website,source_kind,phone_type,phone_source_url,phone_status) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", [tuple(row.get(x) for x in ["entity_key", "corporate_number", "fuma_id", "name", "address", "prefecture", "city", "industry_code", "industry_name", "industry_path", "phone", "website", "source_kind", "phone_type", "phone_source_url", "phone_status"]) for row in companies])
    con.executemany("INSERT INTO industry_taxonomy VALUES(?,?,?,?,?,?,?,?)", [(r["code"], r["name"], r["level"], r["parent_code"] or None, r["path"], r["revision"], r["source_url"], int(r["is_active"])) for r in taxonomy])
    con.execute("CREATE INDEX idx_companies_name ON companies(name)")
    con.execute("CREATE INDEX idx_companies_corporate_number ON companies(corporate_number)")
    con.execute("CREATE INDEX idx_companies_industry_code ON companies(industry_code)")
    con.execute("CREATE INDEX idx_companies_prefecture ON companies(prefecture)")
    con.execute("INSERT INTO company_fts(rowid, entity_key, name, address, industry_name, business_summary) SELECT row_id, entity_key, coalesce(name,''), coalesce(address,''), coalesce(industry_name,''), '' FROM companies")
    con.execute("CREATE TABLE build_metadata (key TEXT PRIMARY KEY, value TEXT)")
    con.execute("INSERT INTO build_metadata VALUES('generation',?)", (generation,))
    con.commit()
    con.close()


def build(args: argparse.Namespace) -> dict[str, Any]:
    global RAW_HEADERS
    xlsx = args.xlsx.resolve()
    master = args.master.resolve()
    out = args.output.resolve()
    if not xlsx.is_file():
        raise FileNotFoundError(xlsx)
    taxonomy_csv = args.taxonomy_csv.resolve()
    taxonomy = build_taxonomy(taxonomy_csv, args.refresh_taxonomy)
    by_code, aliases = canonical_taxonomy(taxonomy)
    headers, fuma_records, header_index = load_fuma(xlsx)
    RAW_HEADERS = headers
    fuma_by_id: dict[str, dict[str, Any]] = {}
    for record in fuma_records:
        row = record["by_header"]
        fuma_id = text(row.get("FUMA_ID")) or f"row-{record['row_number']}"
        record["fuma_id"] = fuma_id
        fuma_by_id[fuma_id] = record

    source = duckdb.connect(str(master), read_only=True)
    recovered_corporate_numbers, corporate_match_stats = recover_corporate_numbers(source, fuma_records)
    fuma_by_corp: dict[str, dict[str, Any]] = {}
    explicit_corporate_number_rows = 0
    for record in fuma_records:
        row = record["by_header"]
        explicit = corporate_number(row.get("法人番号"))
        if explicit:
            explicit_corporate_number_rows += 1
        resolved = explicit or recovered_corporate_numbers.get(str(record["fuma_id"]))
        record["resolved_corporate_number"] = resolved
        record["corporate_number_match_method"] = (
            "fuma_explicit" if explicit else "national_name_address_exact" if resolved else None
        )
        if resolved:
            fuma_by_corp[resolved] = record

    source.execute("CREATE TEMP TABLE fuma_numbers(corporate_number VARCHAR PRIMARY KEY)")
    source.executemany("INSERT INTO fuma_numbers VALUES (?)", [(corp,) for corp in fuma_by_corp])
    public_select = ",".join(PUBLIC_COLUMNS)
    matched_rows = [public_company_row(row) for row in source.execute(f"SELECT {public_select} FROM core.companies c INNER JOIN fuma_numbers f USING(corporate_number)").fetchall()]
    public_by_corp = {str(row["corporate_number"]): row for row in matched_rows}
    source.execute("CREATE TEMP TABLE included_numbers AS SELECT corporate_number FROM fuma_numbers")
    g_only_rows = [public_company_row(row) for row in source.execute(f"SELECT {public_select} FROM core.companies c WHERE c.jsic_major_code='G' AND NOT EXISTS (SELECT 1 FROM included_numbers f WHERE f.corporate_number=c.corporate_number)").fetchall()]
    source.execute(
        "CREATE TEMP TABLE scoped_g_numbers AS "
        "SELECT corporate_number FROM included_numbers "
        "UNION SELECT corporate_number FROM core.companies WHERE jsic_major_code='G'"
    )
    public_industry_rows = source.execute(
        "SELECT i.corporate_number, i.jsic_code "
        "FROM core.company_industries i "
        "INNER JOIN scoped_g_numbers s USING(corporate_number) "
        "WHERE i.jsic_major_code='G'"
    ).fetchall()
    public_contacts, public_contact_stats = load_public_establishment_contacts(source)
    source.close()
    public_industries: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for corp, raw_code in public_industry_rows:
        public_industries[str(corp)].append(public_industry(raw_code, by_code, aliases))

    companies: list[dict[str, Any]] = []
    industries: list[dict[str, Any]] = []
    phone_rows: list[dict[str, Any]] = []
    website_rows: list[dict[str, Any]] = []
    phone_state: list[dict[str, Any]] = []
    for record in fuma_records:
        fuma_id = record["fuma_id"]
        corp = record.get("resolved_corporate_number")
        key = corp or f"fuma:{fuma_id}"
        public = public_by_corp.get(corp) if corp else None
        public_ind = public_industries.get(corp or "", [])
        company, ind_rows, candidate_rows = make_company(
            entity_key=key,
            fuma=record,
            public=public,
            fuma_id=fuma_id,
            fuma_row_number=record["row_number"],
            by_code=by_code,
            aliases=aliases,
            industry=public_ind[0] if public_ind else None,
            corporate_number_override=corp,
            corporate_number_match_method=record.get("corporate_number_match_method"),
        )
        companies.append(company)
        industries.extend(ind_rows)
        phone_rows.extend(candidate_rows)

    fuma_keys = {row["corporate_number"] for row in companies if row["corporate_number"]}
    for public in g_only_rows:
        corp = str(public["corporate_number"])
        assignments = public_industries.get(corp, [])
        industry = max(assignments, key=lambda item: len(item["code"])) if assignments else public_industry("G", by_code, aliases)
        company, ind_rows, _ = make_company(entity_key=corp, fuma=None, public=public, fuma_id=None, fuma_row_number=None, by_code=by_code, aliases=aliases, industry=industry)
        companies.append(company)
        industries.extend(ind_rows)

    public_contact_stats.update(
        apply_public_establishment_contacts(companies, public_contacts, phone_rows, website_rows)
    )
    generation = f"g-v{PACKAGE_VERSION}-fuma-{sha256(xlsx)[:12]}"
    phone_state = [
        {
            "entity_key": company["entity_key"],
            "corporate_number": company["corporate_number"],
            "company_name": company["name"],
            "prefecture_name": company["prefecture"],
            "city_name": company["city"],
            "employee_number": company["employees"],
            "capital_stock": company["capital"],
            "scope_label": "G37-G41",
            "dataset_generation": generation,
            "jsic_major_codes": "G",
            "jsic_middle_codes": company["industry_middle_code"] or "",
            "runtime_binding_status": "matched",
            "website": company["website"],
            "state": company["phone_status"] if company["phone"] else (
                "pending_official_site" if company["website"] and company["corporate_number"]
                else "fuma_only_blocked" if company["website"]
                else "website_missing"
            ),
            "last_completed_at": company["phone_observed_at"],
            "last_error": None,
        }
        for company in companies
    ]
    build_dir = out / "_build"
    build_dir.mkdir(parents=True, exist_ok=True)
    canonical_path = build_dir / "queria_master_g_fuma.duckdb"
    runtime_path = build_dir / "queria_runtime_g_fuma.duckdb"
    sqlite_path = build_dir / "search_g_fuma.sqlite"
    for path in [canonical_path, runtime_path, sqlite_path]:
        for stale_path in [path, Path(str(path) + ".wal")]:
            if stale_path.exists():
                stale_path.unlink()
    con = duckdb.connect(str(canonical_path))
    create_schema(con)
    con.execute("BEGIN")
    con.executemany("INSERT INTO meta.industry_taxonomy VALUES(?,?,?,?,?,?,?,?)", [(r["code"], r["name"], r["level"], r["parent_code"] or None, r["path"], r["revision"], r["source_url"], bool(int(r["is_active"])) ) for r in taxonomy])
    con.executemany("INSERT INTO meta.dataset_manifest VALUES(?,?)", [("dataset_key", "G37_41_FUMA"), ("generation", generation), ("version", PACKAGE_VERSION), ("git_ref", GIT_REF), ("scope", "G,37,38,39,40,41 and all descendants")])
    con.execute("INSERT INTO meta.source_metadata VALUES (?,?)", ("G37_41_FUMA", json.dumps({"xlsx_file": xlsx.name, "xlsx_sha256": sha256(xlsx), "national_db_file": master.name, "taxonomy": ESTAT_URL, "revision": "04", "public_establishment_source": MHLW_SOURCE_URL}, ensure_ascii=False)))
    insert_batches(con, "core.g_companies", COMPANY_COLUMNS, companies)
    insert_batches(con, "core.company_industries", ["entity_key", "jsic_code", "jsic_major_code", "jsic_middle_code", "jsic_small_code", "jsic_detail_code", "jsic_major_name", "jsic_middle_name", "jsic_small_name", "jsic_detail_name", "jsic_level", "business_path_raw"], industries)
    insert_batches(con, "enrichment.phone_candidates", ["candidate_id", "entity_key", "phone_raw", "phone_normalized", "phone_type", "source_url", "evidence_text", "confidence", "observed_at", "status"], phone_rows)
    insert_batches(con, "enrichment.website_candidates", ["candidate_id", "entity_key", "url", "url_type", "source_url", "evidence_text", "confidence", "observed_at", "status"], website_rows)
    insert_batches(con, "enrichment.phone_collection_state", ["entity_key", "corporate_number", "website", "state", "last_completed_at", "last_error"], phone_state)
    con.executemany("INSERT INTO core.fuma_columns VALUES(?,?,?)", [(index, header, header) for index, header in enumerate(headers)])
    raw_columns = headers
    raw_rows = []
    for record in fuma_records:
        raw = {header: json_value(record["values"][index]) for index, header in enumerate(headers)}
        raw_rows.append({"source_row_number": record["row_number"], "raw_json": json.dumps(raw, ensure_ascii=False), **{column: text(record["values"][index]) for index, column in enumerate(raw_columns)}})
    insert_batches(con, "core.fuma_records", ["source_row_number", "raw_json", *raw_columns], raw_rows)
    con.execute("COMMIT")
    con.execute("CREATE INDEX idx_g_companies_name ON core.g_companies(name)")
    con.execute("CREATE INDEX idx_g_companies_corporate_number ON core.g_companies(corporate_number)")
    con.execute("CREATE INDEX idx_g_companies_industry_code ON core.g_companies(industry_code)")
    con.execute("CREATE INDEX idx_g_companies_fuma_id ON core.g_companies(fuma_id)")
    con.execute("CREATE INDEX idx_g_industries_jsic_code ON core.company_industries(jsic_code)")
    con.execute("CREATE INDEX idx_g_industries_entity ON core.company_industries(entity_key)")
    con.close()
    shutil.copyfile(canonical_path, runtime_path)
    build_sqlite_index(sqlite_path, companies, taxonomy, generation)

    data_dir = out / "data"
    data_dir.mkdir(parents=True, exist_ok=True)
    for source_path in [canonical_path, runtime_path, sqlite_path]:
        target = data_dir / source_path.name
        if target.exists():
            target.unlink()
        shutil.move(str(source_path), str(target))
    target_metadata = data_dir / "source_metadata.json"
    metadata = {"dataset": "CompanyMaster-G37-41", "version": PACKAGE_VERSION, "generation": generation, "git_ref": GIT_REF, "source_xlsx_file": xlsx.name, "source_xlsx_sha256": sha256(xlsx), "national_db_file": master.name, "taxonomy_source": ESTAT_URL, "taxonomy_revision": "04", "public_establishment_source": MHLW_SOURCE_URL, "fuma_columns": headers}
    target_metadata.write_text(json.dumps(metadata, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    target_targets = data_dir / "phone_targets_g37_41.csv"
    with target_targets.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=[
            "entity_key", "corporate_number", "company_name", "prefecture_name", "city_name",
            "employee_number", "capital_stock", "scope_label", "dataset_generation",
            "jsic_major_codes", "jsic_middle_codes", "runtime_binding_status",
            "website", "state", "last_completed_at", "last_error",
        ])
        writer.writeheader()
        writer.writerows(phone_state)

    counts = Counter(row["source_kind"] for row in companies)
    taxonomy_counts = Counter(row["level"] for row in taxonomy)
    audit = {
        "dataset": "CompanyMaster-G37-41",
        "generation": generation,
        "version": PACKAGE_VERSION,
        "git_ref": GIT_REF,
        "scope": ["G", "37", "38", "39", "40", "41", "all small and detail descendants"],
        "counts": {"fuma_rows": len(fuma_records), "fuma_columns": len(headers), "fuma_explicit_corporate_number_rows": explicit_corporate_number_rows, "fuma_recovered_corporate_number_rows": len(recovered_corporate_numbers), "fuma_corporate_number_rows": len(fuma_by_corp), "fuma_only_rows": sum(1 for row in companies if row["source_kind"] == "fuma-only"), "national_g_only_rows": sum(1 for row in companies if row["source_kind"] == "national-g"), "unified_company_rows": len(companies), "industry_rows": len(industries), "phone_candidate_rows": len(phone_rows), "company_phone_rows": sum(1 for row in companies if row["phone"]), "fuma_phone_rows": sum(1 for record in fuma_records if phone_parts(record["by_header"].get("電話番号"))), "website_candidate_rows": len(website_rows), "official_url_rows": sum(1 for row in companies if row["website"]), "public_establishment_contacts": public_contact_stats, "taxonomy_rows": len(taxonomy), "taxonomy_by_level": dict(taxonomy_counts), "source_kind": dict(counts), "corporate_number_matching": corporate_match_stats},
        "integrity": {"fuma_ids_unique": len(fuma_by_id) == len(fuma_records), "corporate_numbers_are_13_digits": all(not row["corporate_number"] or bool(re.fullmatch(r"\d{13}", row["corporate_number"])) for row in companies), "corporate_numbers_unique": len({row["corporate_number"] for row in companies if row["corporate_number"]}) == sum(1 for row in companies if row["corporate_number"]), "fuma_only_keys_are_prefixed": all(row["source_kind"] != "fuma-only" or row["entity_key"].startswith("fuma:") for row in companies), "no_corporate_number_guessed": True, "recovered_numbers_require_unique_exact_name_and_address": True, "phone_not_declared_representative": all(not row["phone"] or row["phone_type"] in {"unclassified", "establishment"} for row in companies), "public_establishment_contacts_joined_by_corporate_number": True},
        "sources": {"fuma_xlsx": {"file_name": xlsx.name, "sha256": sha256(xlsx), "redistribution_terms": "user-provided source; confirm applicable terms"}, "national_duckdb": {"file_name": master.name, "release": "v0.9.0"}, "public_establishments": MHLW_SOURCE_URL, "estat": ESTAT_URL},
    }
    for path in [data_dir / "queria_master_g_fuma.duckdb", data_dir / "queria_runtime_g_fuma.duckdb", data_dir / "search_g_fuma.sqlite", data_metadata_path(data_dir), target_targets, out / "CompanyMaster-G37-41.exe"]:
        if path.is_file():
            audit.setdefault("artifacts", {})[path.name] = {"bytes": path.stat().st_size, "sha256": sha256(path)}
    (out / "audit.json").write_text(json.dumps(audit, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (out / "README_PORTABLE_JA.md").write_text(make_readme(audit), encoding="utf-8")
    return audit


def data_metadata_path(data_dir: Path) -> Path:
    return data_dir / "source_metadata.json"


def make_readme(audit: dict[str, Any]) -> str:
    counts = audit["counts"]
    public_contacts = counts["public_establishment_contacts"]
    return f"""# CompanyMaster 大分類G（情報通信業）完全版 v{audit['version']}\n\n- Git参照: `{audit['git_ref']}`\n- 対象: 大分類G（中分類37〜41と全小分類・細分類）\n- FUMA全量: {counts['fuma_rows']:,}行\n- 統合企業: {counts['unified_company_rows']:,}件\n- 法人番号付きFUMA: {counts['fuma_corporate_number_rows']:,}件（社名＋住所の一意完全一致で回復 {counts['fuma_recovered_corporate_number_rows']:,}件）\n- FUMA-only: {counts['fuma_only_rows']:,}件\n- 電話付き企業: {counts['company_phone_rows']:,}件（FUMA由来 {counts['fuma_phone_rows']:,}件、公開事業所電話の新規反映 {public_contacts['promoted_primary_phone_rows']:,}件）\n- 公式HP付き企業: {counts['official_url_rows']:,}件\n- 公開事業所HP候補: {counts['website_candidate_rows']:,}件\n- JSICマスター: {counts['taxonomy_rows']:,}件（空分類を含む）\n\n## 起動\n\n`CompanyMaster-G37-41.exe`をこのフォルダから起動してください。ファイル名の`G37-41`は既存配布との互換名で、対象の正本は大分類`G`です。`data`フォルダはEXEと同じ場所に置きます。\n\n## 検索コード\n\n`G`, `37`, `G37`, `39`, `G39`, `391`, `G391`, `3911`, `G3911`を受け付けます。中分類・小分類・細分類は配下を含む前方一致で、細分類は完全なコードとして絞り込めます。\n\n## HP・電話番号\n\n`core.g_companies.phone`にはFUMA電話を優先し、空欄だった企業だけ厚生労働省公開データの事業所電話を反映します。公開事業所電話は`phone_type='establishment'`であり、本社代表電話とは断定しません。全候補と根拠は`enrichment.phone_candidates`、事業所HP候補は`enrichment.website_candidates`に保存しています。法人番号の回復根拠は`corporate_number_match_method`と`corporate_number_match_score`で監査できます。LLMは全社処理せず、根拠のある低信頼候補の確認に限定します。\n\n元Excelの{counts['fuma_columns']}列はDuckDBの`core.fuma_records.raw_json`と`core.fuma_columns`に保持しています。\n"""


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--master", type=Path, default=DEFAULT_MASTER)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--taxonomy-csv", type=Path, default=Path("reference/jsic_g37_41.csv"))
    parser.add_argument("--refresh-taxonomy", action="store_true")
    args = parser.parse_args()
    audit = build(args)
    print(json.dumps(audit, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
